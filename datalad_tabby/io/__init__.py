"""Input/output and conversion routines for tabby-format metadata records
"""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import (
    Dict,
    List,
)

from openpyxl import (
    Workbook,
    load_workbook,
)
from openpyxl.worksheet.worksheet import Worksheet


__all__ = ['xlsx2tabby', 'tabby2xlsx', 'load_tabby']


def load_tabby(
    src: Path,
    *,
    single: bool = True,
    jsonld: bool = True,
    recursive: bool = True,
) -> Dict | List:
    """Load a tabby (TSV) record as structured (JSON(-LD)) data

    The record is identified by the table/sheet file path ``src``. This need
    not be the root 'dataset' sheet, but can be any component of the full
    record.

    The ``single`` flag determines whether the record is interpreted as a
    single entity (i.e., JSON object), or many entities (i.e., JSON array of
    (homogeneous) objects).  Depending on the ``single`` flag, either a
    ``dict`` or a ``list`` is returned.

    Other tabby tables/sheets are loaded when ``@tabby-single|many-`` import
    statements are discovered. The corresponding data structures then replace
    the import statement at its location. Setting the ``recursive`` flag to ``False``
    disables table import, which will result in only the record available at the
    ``src`` path being loaded.

    .. todo::

      With the ``jsonld`` flag, a declared or default JSON-LD context or frame
      is inserted and/or applied.

    """
    return (_load_tabby_single if single else _load_tabby_many)(
        src=src,
        jsonld=jsonld,
        recursive=recursive,
        trace=[],
    )


def _load_tabby_single(
    *,
    src: Path,
    jsonld: bool,
    recursive: bool,
    trace: List,
) -> Dict:
    obj = {}
    with src.open(newline='') as tsvfile:
        reader = csv.reader(tsvfile, delimiter='\t')
        # row_id is useful for error reporting
        for row_id, row in enumerate(reader):
            # row is a list of field, with only as many items
            # as this particular row has columns
            if not len(row) or not row[0] or row[0].startswith('#'):
                # skip empty rows, rows with no key, or rows with
                # a comment key
                continue
            key = row[0]
            val = row[1:]
            # cut `val` short and remove trailing empty items
            val = val[:_get_index_after_last_nonempty(val)]
            if not val:
                # skip properties with no value(s)
                continue
            # look for @tabby-... imports in values, and act on them
            val = [
                _resolve_value(
                    v,
                    src,
                    jsonld=jsonld,
                    recursive=recursive,
                    trace=trace,
                )
                for v in val
            ]
            # we do not amend values for keys!
            # another row for an already existing key overwrites
            # we support "sequence" values via multi-column values
            # supporting two ways just adds unnecessary complexity
            obj[key] = val

    # apply any overrides
    obj.update(_build_overrides(src, obj))

    obj = _compact_obj(obj)

    if not jsonld:
        # early exit
        return obj

    # with jsonld==True, looks for a context
    ctx = _get_corresponding_context(src)
    if ctx:
        _assigned_context(obj, ctx)

    return obj


def _get_corresponding_context(src):
    rec_ctx_fpath = _get_record_context_fpath(src)
    sheet_ctx_fpath = _get_corresponding_context_fpath(src)
    # TODO take built-in context instead of empty
    ctx = {}
    for ctx_fpath in (rec_ctx_fpath, sheet_ctx_fpath):
        if ctx_fpath.exists():
            custom_ctx = json.load(ctx_fpath.open())
            # TODO report when redefinitions occur
            ctx.update(custom_ctx)

    return ctx


def _assigned_context(obj: Dict, ctx: Dict):
    if '@context' in obj:
        # TODO report when redefinitions occur
        # TODO in principle a table could declare a context, but this is
        # a theoretical possibility that is neither advertised nor tested
        obj['@context'].update(ctx)  # pragma: no cover
    else:
        obj['@context'] = ctx


def _resolve_value(
    v: str,
    src_sheet_fpath: Path,
    jsonld: bool,
    recursive: bool,
    trace: List,
):
    if not recursive or not v.startswith('@tabby-'):
        return v

    if v.startswith('@tabby-single-'):
        loader = _load_tabby_single
        src = _get_corresponding_sheet_fpath(src_sheet_fpath, v[14:])
    elif v.startswith('@tabby-many-'):
        loader = _load_tabby_many
        src = _get_corresponding_sheet_fpath(src_sheet_fpath, v[12:])
    else:
        # strange, but not enough reason to fail
        return v

    trace = _build_import_trace(src, trace)

    return loader(
        src=src,
        jsonld=jsonld,
        recursive=recursive,
        trace=trace,
    )


def _build_overrides(src: Path, obj: Dict):
    overrides = {}
    ofpath = _get_corresponding_override_fpath(src)
    if not ofpath.exists():
        # we have no overrides
        return overrides
    orspec = json.load(ofpath.open())
    for k in orspec:
        spec = orspec[k]
        s = [
            # interpolate str spec, anything else can pass
            # through as-is
            s.format(**obj) if isinstance(s, str) else s
            for s in (spec if isinstance(spec, list) else [spec])
        ]
        overrides[k] = s
    return overrides


def _get_corresponding_sheet_fpath(fpath: Path, sheet_name: str) -> Path:
    return fpath.parent / \
        f'{_get_tabby_prefix_from_sheet_fpath(fpath)}_{sheet_name}.tsv'


def _get_record_context_fpath(fpath: Path) -> Path:
    prefix = _get_tabby_prefix_from_sheet_fpath(fpath)
    return fpath.parent / f'{prefix}.ctx.jsonld'


def _get_corresponding_context_fpath(fpath: Path) -> Path:
    return fpath.parent / f'{fpath.stem}.ctx.jsonld'


#def _get_corresponding_frame_fpath(fpath: Path) -> Path:
#    return fpath.parent / f'{fpath.stem}.frame.jsonld'


def _get_corresponding_override_fpath(fpath: Path) -> Path:
    return fpath.parent / f'{fpath.stem}.override.json'


def _get_tabby_prefix_from_sheet_fpath(fpath: Path) -> str:
    stem = fpath.stem
    # stem up to, but not including, the last '_'
    return stem[:(-1) * stem[::-1].index('_') - 1]


def _get_index_after_last_nonempty(val: List) -> int:
    for i, v in enumerate(val[::-1]):
        if v:
            return len(val) - i
    return 0


def _build_import_trace(src: Path, trace: List) -> List:
    if src in trace:
        raise RecursionError(
            f'circular import: {src} is (indirectly) referencing itself')

    # TODO if we ever want to go parallel, we should produce an independent
    # list instance here
    trace.append(src)
    return trace


def _load_tabby_many(
    *,
    src: Path,
    jsonld: bool,
    recursive: bool,
    trace: List,
) -> List[Dict]:
    array = list()
    fieldnames = None

    # with jsonld==True, looks for a context
    ctx = _get_corresponding_context(src)

    with src.open(newline='') as tsvfile:
        # we cannot use DictReader -- we need to support identically named
        # columns
        reader = csv.reader(tsvfile, delimiter='\t')
        # row_id is useful for error reporting
        for row_id, row in enumerate(reader):
            # row is a list of field, with only as many items
            # as this particular row has columns
            if not len(row) \
                    or row[0].startswith('#') \
                    or all(v is None for v in row):
                # skip empty rows, rows with no key, or rows with
                # a comment key
                continue
            if fieldnames is None:
                # the first non-ignored row defines the property names/keys
                # cut `val` short and remove trailing empty items
                fieldnames = row[:_get_index_after_last_nonempty(row)]
                continue

            obj = _manyrow2obj(src, row, jsonld, fieldnames, recursive, trace)

            if ctx:
                _assigned_context(obj, ctx)

            # simplify single-item lists to a plain value
            array.append(_compact_obj(obj))
    return array


def _manyrow2obj(
    src: Path,
    row: List,
    jsonld: bool,
    fieldnames: List,
    recursive: bool,
    trace: List,
) -> Dict:
    # if we get here, this is a value row, representing an individual
    # object
    obj = {}
    vals = [
        # look for @tabby-... imports in values, and act on them.
        # keep empty for now to maintain fieldname association
        _resolve_value(
            v,
            src,
            jsonld=jsonld,
            recursive=recursive,
            trace=trace,
        ) if v else v
        for v in row
    ]
    if len(vals) > len(fieldnames):
        # we have extra values, merge then into the column
        # corresponding to the last key
        last_key_idx = len(fieldnames) - 1
        lc_vals = vals[last_key_idx:]
        lc_vals = lc_vals[:_get_index_after_last_nonempty(lc_vals)]
        vals[last_key_idx] = lc_vals

    # merge values with keys, amending duplicate keys as necessary
    for i, k in enumerate(fieldnames):
        if i >= len(vals):
            # no more values defined in this row, skip this key
            continue
        v = vals[i]
        if not v:
            # no value, nothing to store or append
            continue
        # treat any key as a potential multi-value scenario
        k_vals = obj.get(k, [])
        k_vals.append(v)
        obj[k] = k_vals

    # TODO optimize and not read spec from file for each row
    # apply any overrides
    obj.update(_build_overrides(src, obj))
    return obj


def _compact_obj(obj: Dict) -> Dict:
    # simplify single-item lists to a plain value
    return {
        k:
        # we let any @context pass-through as-is to avoid ANY sideeffects
        # of fiddling with the complex semantics of context declaration
        # structure
        vals if k == '@context' else vals if len(vals) > 1 else vals[0]
        for k, vals in obj.items()
    }


def xlsx2tabby(src: Path, dest: Path) -> List[Path]:
    """Convert a tabby record in a (multi-sheet) XLSX file to a series of TSVs

    The TSV file collection will use the name of the XLSX file (without
    extension) as name prefix and append ``_<sheetname>.tsv`` it to build the
    full output filename in the ``dest`` directory.
    """
    tabby_prefix = src.stem
    wb = load_workbook(
        filename=src,
        # see https://openpyxl.readthedocs.io/optimized.html#read-only-mode
        read_only=True,
    )
    # TODO we could add support for handling single-sheet workbooks regardless
    # of the name of the sheet (force is to be 'dataset')
    outfpaths = []
    for sheet in wb.sheetnames:
        outfpath = dest / f"{tabby_prefix}_{sheet}.tsv"
        _sheet2tsv(wb[sheet], outfpath)
        outfpaths.append(outfpath)

    return outfpaths


def tabby2xlsx(src: Path, dest: Path) -> Path:
    """Convert a tabby record as a series of TSVs to a (multi-sheet) XLSX file

    The XLSX file will use the name of the TSV collection (without
    extension and the sheet name) and append ``.xslx`` to build the
    full output filename in the ``dest`` directory.
    """
    tabby_prefix = src.stem
    if tabby_prefix.endswith('_dataset'):
        # this is the root sheet, strip sheet name to discover all
        # other components
        tabby_prefix = tabby_prefix[:-8]

    sheets_paths = list(src.parent.glob(f'{tabby_prefix}_*.tsv'))
    sheet_names = [
        # the sheet name is the last file name component ('_' delimiter)
        # without the '.tsv' extension
        p.name.split('_')[-1][:-4]
        for p in sheets_paths
    ]
    if 'dataset' not in sheet_names:
        raise ValueError(
            f"'dataset' sheet not found for tabby record {tabby_prefix!r}")

    wb = Workbook(
        # https://openpyxl.readthedocs.io/optimized.html#write-only-mode
        write_only=True,
        # prefer ISO 8601 for less interpretation freedom of dates
        # see https://openpyxl.readthedocs.io/en/stable/datetime.html
        iso_dates=True,
    )

    for sheet_name, sheet_path in sorted(zip(sheet_names, sheets_paths)):
        ws = wb.create_sheet(
            sheet_name,
            # put the 'dataset' sheet first, and the rest in sorted order
            # after it
            0 if sheet_name == 'dataset' else None,
        )
        with sheet_path.open(newline='') as tsvfile:
            reader = csv.reader(tsvfile, delimiter='\t')
            for row in reader:
                ws.append(row)

    outfpath = dest / f'{tabby_prefix}.xlsx'
    wb.save(outfpath)
    return outfpath


def _sheet2tsv(ws: Worksheet, dest: Path):
    with dest.open('w', newline='') as tsvfile:
        writer = csv.writer(
            tsvfile,
            delimiter='\t',
        )
        writer.writerows(ws.iter_rows(values_only=True))
