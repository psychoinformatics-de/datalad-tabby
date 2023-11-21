"""Conversion of `tabby` TSV files to and from XLSX multi-sheet spreadsheets
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import List

from openpyxl import (
    Workbook,
    load_workbook,
)
from openpyxl.worksheet.worksheet import Worksheet

__all__ = ['xlsx2tabby', 'tabby2xlsx']


def xlsx2tabby(src: Path, dest: Path) -> List[Path]:
    """Convert a tabby record in a (multi-sheet) XLSX file to a series of TSVs

    The TSV file collection will use the name of the XLSX file (without
    extension) as name prefix and append ``_<sheetname>.tsv`` it to build the
    full output filename in the ``dest`` directory.
    """
    tabby_prefix = src.stem
    wb = load_workbook(
        filename=src,
        # see https://openpyxl.readthedocs.io/optimized.html#read-only-mode
        read_only=True,
    )
    # TODO we could add support for handling single-sheet workbooks regardless
    # of the name of the sheet (force is to be 'dataset')
    outfpaths = []
    for sheet in wb.sheetnames:
        outfpath = dest / f"{tabby_prefix}_{sheet}.tsv"
        _sheet2tsv(wb[sheet], outfpath)
        outfpaths.append(outfpath)

    return outfpaths


def tabby2xlsx(src: Path, dest: Path) -> Path:
    """Convert a tabby record as a series of TSVs to a (multi-sheet) XLSX file

    The XLSX file will use the name of the TSV collection (without
    extension and the sheet name) and append ``.xslx`` to build the
    full output filename in the ``dest`` directory.
    """
    tabby_prefix = src.stem
    if tabby_prefix.endswith('_dataset'):
        # this is the root sheet, strip sheet name to discover all
        # other components
        tabby_prefix = tabby_prefix[:-8]

    sheets_paths = list(src.parent.glob(f'{tabby_prefix}_*.tsv'))
    sheet_names = [
        # the sheet name is the last file name component ('_' delimiter)
        # without the '.tsv' extension
        p.name.split('_')[-1][:-4]
        for p in sheets_paths
    ]
    if 'dataset' not in sheet_names:
        raise ValueError(
            f"'dataset' sheet not found for tabby record {tabby_prefix!r}")

    wb = Workbook(
        # https://openpyxl.readthedocs.io/optimized.html#write-only-mode
        write_only=True,
        # prefer ISO 8601 for less interpretation freedom of dates
        # see https://openpyxl.readthedocs.io/en/stable/datetime.html
        iso_dates=True,
    )

    for sheet_name, sheet_path in sorted(zip(sheet_names, sheets_paths)):
        ws = wb.create_sheet(
            sheet_name,
            # put the 'dataset' sheet first, and the rest in sorted order
            # after it
            0 if sheet_name == 'dataset' else None,
        )
        with sheet_path.open(newline='') as tsvfile:
            reader = csv.reader(tsvfile, delimiter='\t')
            for row in reader:
                ws.append(row)

    outfpath = dest / f'{tabby_prefix}.xlsx'
    wb.save(outfpath)
    return outfpath


def _sheet2tsv(ws: Worksheet, dest: Path):
    with dest.open('w', newline='') as tsvfile:
        writer = csv.writer(
            tsvfile,
            delimiter='\t',
        )

        # find the last nonempty row
        max_idx = 1
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if any(v is not None for v in row):
                max_idx = i + 1  # max row is a 1-based index

        # write tsv, truncating empty rows at the end
        writer.writerows(ws.iter_rows(values_only=True, max_row=max_idx))

